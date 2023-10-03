import openpyxl
import re
import time
import os
import json
from numpy import sin
import pandas as pd
import json
import math
import random

loc = "data.xlsx"

termID_docID = []
inverted_index = {}
tfIdf_values = {}
championsList = {}
stop_words = ["از", "در", "به", "با", "به", "تا", "را", "و", "بی", "مگر", "الا", "اندر", "غیر", "است", "این", "می",
              "برای", "انتهای", "پیام", "ایسنا", "ها", "تر", "ترین", "ها", "های", "ات", "آسا", "آگین", "که", "وار"]
suffix_list = ["تر", "ترین", "ها", "های", "ات", "آسا", "آگین", "وار"]
prefix_list = ["ابر", "اندر", "باز", "بی", "پاد", "فرا", "می", "نا"]
verbs = ["دار", "گفت", "هست", "کرد" "رفت", "اورد", "بخشید", "برخاست", "گزاشت", "گذاشت", "خواه", "ریخت", "پوشید"
                                                                                                        "توان",
         "گرفت" "ساخت", "دانست", "گرفت", "دید", "آمد", "امد", "خواست", "داد", "آورد", "بود", "برگشت", "گشت"]


# this method will prune the words and remove commas and dots and so on ...
# for example: سلام، --> سلام
def pruning_word(words):
    pruned_words = [
        x.replace('.', '').replace(':', '').replace('،', '').replace(';', '').replace('*', '').replace('+', '')
            .replace('**', '').replace('؛', '').replace(']', '').replace('[', '').replace('\"', '')
            .replace('»', '').replace('«', '').replace('?', '').replace('!', '').replace('#', '')
            .replace('^', '').replace('&', '').replace('(', '').replace(')', '').replace('؟', '')
            .replace('“', '').replace('–', '').replace('-', '').replace('_', '').replace('%', '')
            .replace('0', '').replace('1', '').replace('2', '').replace('3', '').replace('4', '')
            .replace('5', '').replace('6', '').replace('7', '').replace('8', '').replace('9', '')
            .replace('۰', '').replace('۱', '').replace('۲', '').replace('۳', '').replace('۴', '')
            .replace('۵', '').replace('۶', '').replace('۷', '').replace('۸', '').replace('۹', '')
            .replace('\u200c', '').replace('\u202b', '').replace('\u200f', '') for x in words]

    term_to_remove = []
    for i in range(len(pruned_words)):
        if ('https' in pruned_words[i]) or (len(pruned_words[i]) >= 20) or (pruned_words[i].isdigit()) \
                or (pruned_words[i] == '+') or (len(pruned_words[i]) == 0) or (pruned_words[i] == '…') \
                or (pruned_words[i] == '–'):
            term_to_remove.append(pruned_words[i])
    for i in range(len(term_to_remove)):
        pruned_words.remove(term_to_remove[i])
    # print(term_to_remove)
    return pruned_words


def tokenizer(text):
    words = pruning_word(text.split())
    # words_without_duplicate = []
    # [words_without_duplicate.append(word) for word in words if word not in words_without_duplicate]
    sorted_set = sorted(words)
    del words
    # del words_without_duplicate
    return sorted_set


def term_docID_maker(tokens, id):
    term_docID = []
    i = 0
    for t in tokens:
        term_docID.append([t, id])
        i += 1
    return term_docID


def term_docID_for_all_doc_maker(sheet):
    rows = sheet.max_row
    columns = sheet.max_column
    docs_term_docID = []  # each document termID_docID term
    for i in range(2, rows + 1):
        id = sheet.cell(i, 1).value
        text = sheet.cell(i, 2).value
        tokens = tokenizer(text)
        docs_term_docID.append(term_docID_maker(tokens, id))
        del text
        del tokens
    return docs_term_docID


def indexer(sheet):
    docs_term_docID = term_docID_for_all_doc_maker(sheet)
    global termID_docID
    for d in docs_term_docID:
        for term in d:
            termID_docID.append(term)
    termID_docID = sorted(termID_docID)
    # for term in termID_docID:
    #     print(term)


def inverted_index_maker():
    terms = []
    for term in termID_docID:
        terms.append(term[0])
    global inverted_index
    inverted_index = dict.fromkeys(terms, "posting list")
    pointer = 0
    end_border = len(termID_docID)
    IS_FINISHED = False
    # print("end border: ", end_border)
    for key, value in inverted_index.items():
        posting = []
        # print("key : " + key)
        while key == termID_docID[pointer][0]:
            # print("pointer = ", pointer, " term = " + termID_docID[pointer][0])
            posting.append(termID_docID[pointer][1])
            pointer += 1
            if pointer == end_border:
                IS_FINISHED = True
                break
        inverted_index[key] = posting
        del posting
        if IS_FINISHED:
            break


def delete_stop_words():
    global stop_words
    for stop_word in stop_words:
        if stop_word in inverted_index:
            del inverted_index[stop_word]


def merge_postings(new_posting, old_posting):
    new_posting.extend(old_posting)
    new_posting.sort()
    return new_posting


def delete_prefix():
    global prefix_list
    terms_to_modify = {}
    for prefix in prefix_list:
        for key, value in inverted_index.items():
            if key.startwith(prefix):
                terms_to_modify[key] = prefix

    for key, value in terms_to_modify.items():
        # print(key, value)
        posting = inverted_index[key]
        del inverted_index[key]
        res = re.sub(value, '', key)
        # print("res: ", res)
        if res in inverted_index:
            # print("word existed before: ", res)
            posting = merge_postings(posting, inverted_index[res])
        inverted_index[res] = posting


def delete_suffix():
    global suffix_list
    terms_to_modify = {}
    for suffix in suffix_list:
        for key, value in inverted_index.items():
            if key.endswith(suffix):
                terms_to_modify[key] = suffix

    for key, value in terms_to_modify.items():
        # print(key, value)
        posting = inverted_index[key]
        del inverted_index[key]
        res = re.sub(value, '', key)
        # print("res: ", res)
        if res in inverted_index:
            # print("word existed before: ", res)
            posting = merge_postings(posting, inverted_index[res])
        inverted_index[res] = posting


def stemming():
    global verbs
    terms_to_modify = {}
    for verb in verbs:
        for key, value in inverted_index.items():
            if verb in key and not key.endswith(verb):
                terms_to_modify[key] = verb
    for key, value in terms_to_modify.items():
        # print(key, value)
        posting = inverted_index[key]
        del inverted_index[key]
        if value in inverted_index.keys():
            # print("word existed before: ", res)
            posting = merge_postings(posting, inverted_index[value])
        inverted_index[value] = posting


def indexing(sheet):
    indexer(sheet)
    inverted_index_maker()
    delete_stop_words()
    # for k, v in inverted_index.items():
    #     print(k, v)
    # delete_suffix()
    # delete_prefix()
    # stemming()


# TODO: retuen URL of the result
def search_query(query):
    # print(query, len(query))
    result = []
    if len(query) == 1:
        if query[0] in inverted_index:
            result = inverted_index[query[0]]
    if len(query) > 1:
        postings_list = []
        for word in query:
            if word in inverted_index:
                postings_list.append(inverted_index[word])
            else:
                postings_list.append([])
        # for p in range(len(postings_list)):
        #     print(query[p], postings_list[p])
        if len(query) == 2:
            # print("debug 0")
            if not len(postings_list[0]) == 0 and not len(postings_list[1]) == 0:
                # print("debug 1")
                result = list(set(postings_list[0]).intersection(postings_list[1]))
        if len(query) == 3:
            # all words in document
            result = list(set(postings_list[0]).intersection(postings_list[1]))
            result = list(set(result).intersection(postings_list[2]))
            # two word in document
            intersect_0_1 = list(set(postings_list[0]).intersection(postings_list[1]))
            intersect_0_2 = list(set(postings_list[0]).intersection(postings_list[2]))
            intersect_1_2 = list(set(postings_list[1]).intersection(postings_list[2]))
            result.extend(intersect_0_1)
            result.extend(intersect_0_2)
            result.extend(intersect_1_2)
    return list(set(result))


def saveData(data, file_name):
    if os.path.exists(file_name):
        os.remove(file_name)
    with open(file_name, 'w', encoding="utf-8") as convert_file:
        convert_file.write(json.dumps(data, ensure_ascii=False))


def readData(file_name):
    with open(file_name, encoding="utf-8") as json_file:
        return json.load(json_file)


def tf(repetition):
    return 1 + math.log(repetition, 10)


def idf(term, n):
    if term not in stop_words:
        if term not in inverted_index:
            return 0
        else:
            term_posting_list = inverted_index[term]
            df = len(set(term_posting_list))
    else:
        df = n
    return math.log(n / df, 10)


# output of this method is a nested dictionary
# the keys of the dictionary is document id's
# value of each key is another dictionary that contains tf_idf value of term t in doc d
# this dictionary keys are terms of a dictionary and value of each key is tf_idf value
# HOW DOES ALGORITHM WORK:
    # tf and idf of each term t in document d is needed to calculate tf_idf value
    # tf is frequency of term t in document d, so in order to calculate that we need to now how many times t repeated in d
        # the main part of this method is written to calculate tf(term frequency of term t in document d)
    # idf is document frequency of term t in all docs, in order to calculate that we can use inverted_index that we create before
    # get the posting list of term t from inverted index and use size of the posting list to calculate idf
def tfIdf(sheet):
    docs_term_docID = term_docID_for_all_doc_maker(sheet)
    # print("number of docs: ", len(docs_term_docID))
    global tfIdf_values
    doc_number = 0
    for doc in docs_term_docID:
        term_repetition_counter = 1
        upper_bound = len(doc) - 1
        # print("upper_bound: ", upper_bound)
        pointer = 0
        tfIdfs_of_doc_d = {}
        while True:
            term = doc[pointer][0]
            # print("pointer= ", pointer, " term: ", term)
            tf_value = 0
            is_tf_found = False
            is_repeated = False
            is_last_term_not_duplicated = False
            for j in range(pointer + 1, upper_bound + 1):
                # print("j: ", j, doc[j][0])
                if term == doc[j][0]:
                    term_repetition_counter += 1
                    is_repeated = True
                else:
                    _tf = tf(term_repetition_counter)
                    term_repetition_counter = 1
                    is_tf_found = True
                    pointer = j
                    if j == upper_bound and doc[j][0] != term:
                        is_last_term_not_duplicated = True
                    break
                if j == upper_bound:
                    if is_repeated:
                        _tf = tf(term_repetition_counter)
                        is_tf_found = True
                    pointer = upper_bound
                    break
            if is_tf_found:
                _idf = idf(term, len(docs_term_docID))
            tfIdfs_of_doc_d[term] = _tf * _idf
            if pointer == upper_bound:
                # print("debug 0")
                if is_last_term_not_duplicated:
                    _tf = tf(term_repetition_counter)
                    _idf = idf(term, len(docs_term_docID))
                    tfIdfs_of_doc_d[doc[upper_bound][0]] = _tf * _idf
                    is_last_term_not_duplicated = False
                break
        tfIdf_values[doc_number] = tfIdfs_of_doc_d
        doc_number += 1
    # for k, v in tfIdf_values.items():
    #     print(k)
    #     for kk, vv in v.items():
    #         print(kk, vv)
    #     print("******")


def champion_list():
    k = 10
    global championsList
    for term, posting in inverted_index.items():
        # print(term, posting)
        distinct_posting = set(posting)
        score = {}
        for post in distinct_posting:
            score[post] = posting.count(term)
        if len(distinct_posting) > k:
            score = dict(sorted(score.items(), key=lambda item: item[1], reverse=True)[0:k])
        else:
            score = dict(sorted(score.items(), key=lambda item: item[1], reverse=True))
        championsList[term] = list(score.keys())


def cosine_similarity(a, b):
    # print("size query: ", len(a))
    # for k in b.keys():
    #     print(k)
    sigma = 0
    for term, wight in a.items():
        # print(term, wight)
        if term in b.keys():
            sigma += wight * b[term]
        else:
            sigma += 0
    # print("sigma: ", sigma)
    l2 = 0
    for term, wight in b.items():
        l2 += math.pow(wight, 2)
    l2 = math.sqrt(l2)
    # print("l2", l2)
    return sigma/l2


def naive_query_searching(query_tfIdf_value):
    similarity = {}
    for k,v in tfIdf_values.items():
        similarity[k] = cosine_similarity(query_tfIdf_value, v)
    # similarity = {k: v for k, v in sorted(similarity.items(), key=lambda item: item[1], reverse=True)}
    return similarity
    # for k,v in similarity.items():
    #     print(k, v)


def index_elimination(query_tfIdf_value, is_championList_used):
    postings = []
    for key in query_tfIdf_value.keys():
        if is_championList_used:
            if key in championsList:
                postings.extend(set(championsList[key]))
        else:
            if key in inverted_index:
                postings.extend(set(inverted_index[key]))
    postings = sorted(set(postings))
    # print("merged")
    # print(postings)
    similarity = {}
    for post in postings:
        similarity[post] = cosine_similarity(query_tfIdf_value, tfIdf_values[str(post-1)])
    # similarity = {k: v for k, v in sorted(similarity.items(), key=lambda item: item[1], reverse=True)}
    # for k,v in similarity.items():
    #     print(k, v)
    return similarity


def query_processing(query):
    is_championList = True
    distinct_term = set(query)
    query_tfIdf_value = {}
    n = len(tfIdf_values)
    for term in distinct_term:
        query_tfIdf_value[term] = tf(query.count(term)) * idf(term, n)
    similarity = naive_query_searching(query_tfIdf_value)
    # similarity = index_elimination(query_tfIdf_value, is_championList)
    k = 10
    similarity = dict(sorted(similarity.items(), key=lambda item: item[1], reverse=True)[0:k])
    for k,v in similarity.items():
        print(k, v)


def main():
    mod = int(input("1: indexing data file from scratch \n2: using extracted data "))

    wb = openpyxl.load_workbook(loc)
    sheet = wb['Sheet1']
    global inverted_index
    global tfIdf_values
    global championsList
    print(mod)
    if mod == 1:  # create inverted-index & tfIdf-values from data and save the result in files
        print("mod == 1")
        indexing(sheet)
        saveData(inverted_index, "inverted_index.txt")
        tfIdf(sheet)
        saveData(tfIdf_values, "tf_idf_values.txt")
        champion_list()
        saveData(championsList, "champions_list.txt")
    elif mod == 2:  # read data from inverted-index & tfIdf-values
        print("mod == 2")
        inverted_index = readData("inverted_index.txt")
        tfIdf_values = readData("tf_idf_values.txt")
        championsList = readData("champions_list.txt")
    else:
        print("wrong input")
        return

    while True:
        query = input("query: ").split()
        if query[0] == 'q':
            break
        start_time = time.time()
        query_processing(query)
        end_time = time.time()
        print("time: ", end_time - start_time)
        # if len(result) == 0:
        #     print("no result for query: ", " ".join(query))
        # else:
        #     print(result)


main()
